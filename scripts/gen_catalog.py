"""
Generate realistic product catalog XLS for RAG testing.
5000 rows: smartphones, laptops, smartwatches, tablets.
One row per variant. Columns match the import UI field mapping.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import itertools, random, math

random.seed(42)

# ─── SMARTPHONES ──────────────────────────────────────────────────────────────
SMARTPHONE_MODELS = [
    # Apple iPhone
    {"brand": "Apple", "series": "iPhone 16 Pro Max", "chipset": "A18 Pro", "camera": "48MP Triple", "battery": "4685mAh", "base_price": 22_000_000, "ram_storage": [("8GB","256GB",0), ("8GB","512GB",2_000_000), ("8GB","1TB",4_000_000)], "colors": ["Black Titanium","White Titanium","Desert Titanium","Natural Titanium"]},
    {"brand": "Apple", "series": "iPhone 16 Pro", "chipset": "A18 Pro", "camera": "48MP Triple", "battery": "3582mAh", "base_price": 19_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",1_500_000), ("8GB","512GB",3_000_000)], "colors": ["Black Titanium","White Titanium","Desert Titanium","Natural Titanium"]},
    {"brand": "Apple", "series": "iPhone 16 Plus", "chipset": "A18", "camera": "48MP Dual", "battery": "4674mAh", "base_price": 16_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",1_500_000), ("8GB","512GB",3_000_000)], "colors": ["Ultramarine","Teal","Pink","White","Black"]},
    {"brand": "Apple", "series": "iPhone 16", "chipset": "A18", "camera": "48MP Dual", "battery": "3561mAh", "base_price": 14_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",1_200_000), ("8GB","512GB",2_500_000)], "colors": ["Ultramarine","Teal","Pink","White","Black"]},
    {"brand": "Apple", "series": "iPhone 15", "chipset": "A16 Bionic", "camera": "48MP Dual", "battery": "3349mAh", "base_price": 11_500_000, "ram_storage": [("6GB","128GB",0), ("6GB","256GB",1_200_000), ("6GB","512GB",2_500_000)], "colors": ["Black","Blue","Green","Yellow","Pink"]},
    {"brand": "Apple", "series": "iPhone 15 Pro", "chipset": "A17 Pro", "camera": "48MP Triple", "battery": "3274mAh", "base_price": 17_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",1_500_000), ("8GB","512GB",3_000_000), ("8GB","1TB",5_000_000)], "colors": ["Black Titanium","White Titanium","Blue Titanium","Natural Titanium"]},
    {"brand": "Apple", "series": "iPhone 14", "chipset": "A15 Bionic", "camera": "12MP Dual", "battery": "3279mAh", "base_price": 9_000_000, "ram_storage": [("6GB","128GB",0), ("6GB","256GB",1_000_000), ("6GB","512GB",2_000_000)], "colors": ["Midnight","Starlight","Blue","Purple","Red"]},
    {"brand": "Apple", "series": "iPhone 13", "chipset": "A15 Bionic", "camera": "12MP Dual", "battery": "3227mAh", "base_price": 7_500_000, "ram_storage": [("4GB","128GB",0), ("4GB","256GB",900_000), ("4GB","512GB",1_800_000)], "colors": ["Midnight","Starlight","Blue","Pink","Red","Green"]},
    # Samsung
    {"brand": "Samsung", "series": "Galaxy S25 Ultra", "chipset": "Snapdragon 8 Elite", "camera": "200MP Quad", "battery": "5000mAh", "base_price": 21_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",2_000_000), ("12GB","1TB",4_000_000)], "colors": ["Titanium Silverblue","Titanium Black","Titanium White Silver","Titanium Jade"]},
    {"brand": "Samsung", "series": "Galaxy S25+", "chipset": "Snapdragon 8 Elite", "camera": "50MP Triple", "battery": "4900mAh", "base_price": 16_500_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",2_000_000)], "colors": ["Icyblue","Coralred","Mint","Silver Shadow"]},
    {"brand": "Samsung", "series": "Galaxy S25", "chipset": "Snapdragon 8 Elite", "camera": "50MP Triple", "battery": "4000mAh", "base_price": 13_500_000, "ram_storage": [("12GB","128GB",0), ("12GB","256GB",1_500_000)], "colors": ["Icyblue","Coralred","Mint","Navy","Silver Shadow"]},
    {"brand": "Samsung", "series": "Galaxy S24 Ultra", "chipset": "Snapdragon 8 Gen 3", "camera": "200MP Quad", "battery": "5000mAh", "base_price": 19_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",2_000_000), ("12GB","1TB",4_000_000)], "colors": ["Titanium Black","Titanium Gray","Titanium Violet","Titanium Yellow"]},
    {"brand": "Samsung", "series": "Galaxy A55 5G", "chipset": "Exynos 1480", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 5_500_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",800_000), ("12GB","256GB",1_200_000)], "colors": ["Awesome Iceblue","Awesome Lilac","Awesome Navy","Awesome Lemon"]},
    {"brand": "Samsung", "series": "Galaxy A35 5G", "chipset": "Exynos 1380", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 4_000_000, "ram_storage": [("6GB","128GB",0), ("8GB","128GB",500_000), ("8GB","256GB",1_000_000)], "colors": ["Awesome Iceblue","Awesome Lilac","Awesome Navy"]},
    {"brand": "Samsung", "series": "Galaxy A25 5G", "chipset": "Exynos 1280", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 3_000_000, "ram_storage": [("6GB","128GB",0), ("8GB","128GB",400_000), ("8GB","256GB",800_000)], "colors": ["Blue","Black","Yellow"]},
    {"brand": "Samsung", "series": "Galaxy M35 5G", "chipset": "Exynos 1380", "camera": "50MP Triple", "battery": "6000mAh", "base_price": 3_500_000, "ram_storage": [("6GB","128GB",0), ("8GB","128GB",400_000), ("8GB","256GB",800_000)], "colors": ["Thunder Gray","Midnight Blue","Coral Blue"]},
    # Xiaomi
    {"brand": "Xiaomi", "series": "14 Ultra", "chipset": "Snapdragon 8 Gen 3", "camera": "50MP Leica Quad", "battery": "5000mAh", "base_price": 17_500_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000)], "colors": ["Black","White","Titanium Special Edition"]},
    {"brand": "Xiaomi", "series": "14T Pro", "chipset": "Dimensity 9300+", "camera": "50MP Leica Triple", "battery": "5000mAh", "base_price": 10_500_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_500_000), ("16GB","512GB",2_000_000)], "colors": ["Titan Black","Titan Gray","Titan Blue"]},
    {"brand": "Xiaomi", "series": "14T", "chipset": "Dimensity 8300-Ultra", "camera": "50MP Leica Triple", "battery": "5000mAh", "base_price": 7_500_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_200_000)], "colors": ["Titan Black","Titan Gray","Lemon Green"]},
    {"brand": "Xiaomi", "series": "Redmi Note 13 Pro+", "chipset": "Dimensity 7200-Ultra", "camera": "200MP Triple", "battery": "4600mAh", "base_price": 4_500_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",500_000), ("12GB","512GB",1_000_000)], "colors": ["Aurora Purple","Moonlight White","Midnight Black","Fusion White"]},
    {"brand": "Xiaomi", "series": "Redmi Note 13 Pro", "chipset": "Snapdragon 7s Gen 2", "camera": "200MP Triple", "battery": "5100mAh", "base_price": 3_500_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",500_000)], "colors": ["Arctic White","Ocean Teal","Midnight Black","Lavender Purple"]},
    {"brand": "Xiaomi", "series": "Redmi Note 13", "chipset": "Snapdragon 685", "camera": "108MP Triple", "battery": "5000mAh", "base_price": 2_500_000, "ram_storage": [("6GB","128GB",0), ("8GB","128GB",300_000), ("8GB","256GB",600_000)], "colors": ["Arctic White","Ocean Teal","Midnight Black","Lavender Purple"]},
    {"brand": "Xiaomi", "series": "Redmi 13C", "chipset": "Helio G85", "camera": "50MP Dual", "battery": "5000mAh", "base_price": 1_600_000, "ram_storage": [("4GB","128GB",0), ("6GB","128GB",200_000), ("8GB","256GB",500_000)], "colors": ["Clover Green","Starlight Black","Glacier White"]},
    {"brand": "Xiaomi", "series": "POCO X6 Pro", "chipset": "Dimensity 8300-Ultra", "camera": "64MP Triple", "battery": "5000mAh", "base_price": 5_000_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",500_000), ("12GB","512GB",1_000_000)], "colors": ["Black","Yellow","Grey"]},
    {"brand": "Xiaomi", "series": "POCO M6 Pro", "chipset": "Helio G99-Ultra", "camera": "64MP Triple", "battery": "5000mAh", "base_price": 2_800_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",400_000)], "colors": ["Black","Blue","Purple"]},
    # Realme
    {"brand": "Realme", "series": "GT 6", "chipset": "Snapdragon 8s Gen 3", "camera": "50MP Triple", "battery": "5500mAh", "base_price": 6_500_000, "ram_storage": [("12GB","256GB",0), ("16GB","512GB",1_500_000)], "colors": ["Fluid Silver","Razor Green"]},
    {"brand": "Realme", "series": "12 Pro+", "chipset": "Snapdragon 7s Gen 2", "camera": "50MP Periscope Triple", "battery": "5000mAh", "base_price": 5_000_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",500_000), ("12GB","512GB",1_000_000)], "colors": ["Navigator Beige","Submarine Blue","Coral Purple"]},
    {"brand": "Realme", "series": "12 Pro", "chipset": "Snapdragon 6 Gen 1", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 3_800_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",400_000)], "colors": ["Navigator Beige","Submarine Blue"]},
    {"brand": "Realme", "series": "C67 4G", "chipset": "Snapdragon 685", "camera": "108MP Dual", "battery": "5000mAh", "base_price": 2_200_000, "ram_storage": [("8GB","256GB",0)], "colors": ["Black Rock","Sunny Oasis","Dark Purple"]},
    {"brand": "Realme", "series": "C55", "chipset": "Helio G88", "camera": "64MP Dual", "battery": "5000mAh", "base_price": 1_900_000, "ram_storage": [("6GB","128GB",0), ("8GB","256GB",300_000)], "colors": ["Rainforest","Sunshower"]},
    {"brand": "Realme", "series": "Narzo 70 Pro", "chipset": "Dimensity 7050", "camera": "50MP Dual", "battery": "5000mAh", "base_price": 2_800_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",400_000)], "colors": ["Glass Gold","Glass Black"]},
    # OPPO
    {"brand": "OPPO", "series": "Find X8 Pro", "chipset": "Dimensity 9400", "camera": "50MP Hasselblad Quad", "battery": "5910mAh", "base_price": 16_000_000, "ram_storage": [("16GB","256GB",0), ("16GB","512GB",2_000_000)], "colors": ["Space Black","Pearl White"]},
    {"brand": "OPPO", "series": "Find X8", "chipset": "Dimensity 9400", "camera": "50MP Hasselblad Triple", "battery": "5630mAh", "base_price": 12_000_000, "ram_storage": [("12GB","256GB",0), ("16GB","256GB",1_500_000), ("16GB","512GB",2_500_000)], "colors": ["Space Black","Pearl White","Starry Blue"]},
    {"brand": "OPPO", "series": "Reno 12 Pro", "chipset": "Dimensity 9200+", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 7_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_000_000)], "colors": ["Nebula Silver","Sunset Peach","Space Brown"]},
    {"brand": "OPPO", "series": "Reno 12", "chipset": "Dimensity 9200+", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 5_500_000, "ram_storage": [("12GB","256GB",0)], "colors": ["Astro Silver","Sunset Peach","Matte Brown"]},
    {"brand": "OPPO", "series": "A3 Pro 5G", "chipset": "Dimensity 7300", "camera": "64MP Dual", "battery": "5100mAh", "base_price": 3_500_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",400_000)], "colors": ["Starlight White","Starry Black","Pink"]},
    {"brand": "OPPO", "series": "A60", "chipset": "Snapdragon 695", "camera": "50MP Dual", "battery": "5000mAh", "base_price": 2_600_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",350_000)], "colors": ["Starry Black","Ripple Blue","Purple"]},
    # Vivo
    {"brand": "Vivo", "series": "X200 Pro", "chipset": "Dimensity 9400", "camera": "50MP Zeiss Quad", "battery": "5800mAh", "base_price": 14_500_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000)], "colors": ["Titanium Black","Titanium Gray"]},
    {"brand": "Vivo", "series": "X200", "chipset": "Dimensity 9400", "camera": "50MP Zeiss Triple", "battery": "5800mAh", "base_price": 10_000_000, "ram_storage": [("12GB","256GB",0), ("16GB","256GB",1_500_000), ("16GB","512GB",2_500_000)], "colors": ["Cosmos Black","Titanium Gray","Breeze Green"]},
    {"brand": "Vivo", "series": "V40 Pro", "chipset": "Snapdragon 7 Gen 3", "camera": "50MP Zeiss Triple", "battery": "5500mAh", "base_price": 7_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_000_000)], "colors": ["Titanium Gray","Lotus Purple"]},
    {"brand": "Vivo", "series": "V40", "chipset": "Snapdragon 7 Gen 3", "camera": "50MP Zeiss Triple", "battery": "5500mAh", "base_price": 5_500_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",700_000)], "colors": ["Ganges Blue","Lotus Purple"]},
    {"brand": "Vivo", "series": "Y300 Pro", "chipset": "Snapdragon 7 Gen 3", "camera": "50MP Dual", "battery": "6500mAh", "base_price": 3_800_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",500_000)], "colors": ["Lemon Green","Twilight Purple","Nebula Blue"]},
    {"brand": "Vivo", "series": "Y200 5G", "chipset": "Snapdragon 4 Gen 2", "camera": "50MP Dual", "battery": "5000mAh", "base_price": 2_800_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",350_000)], "colors": ["Sequoia Green","Sahara Dust","Panda White"]},
    # Infinix / Tecno
    {"brand": "Infinix", "series": "Note 40 Pro+", "chipset": "Helio G99 Ultimate", "camera": "108MP Triple", "battery": "4600mAh", "base_price": 3_000_000, "ram_storage": [("12GB","256GB",0), ("16GB","256GB",400_000)], "colors": ["Racing Black","Volcanic Orange","Titan Gold"]},
    {"brand": "Infinix", "series": "Zero 40 5G", "chipset": "Dimensity 8020", "camera": "108MP Triple", "battery": "5000mAh", "base_price": 3_500_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",500_000)], "colors": ["Starfall Silver","Sage Green","Nebula Blue"]},
    {"brand": "Infinix", "series": "Hot 40 Pro", "chipset": "Helio G99", "camera": "108MP Triple", "battery": "5000mAh", "base_price": 2_100_000, "ram_storage": [("8GB","256GB",0)], "colors": ["Volcanic Orange","Starfall Silver","Racing Black"]},
    {"brand": "Tecno", "series": "Phantom V Fold2 5G", "chipset": "Dimensity 9000+", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 8_500_000, "ram_storage": [("12GB","512GB",0)], "colors":["Innovative Black","Innovative White"]},
    {"brand": "Tecno", "series": "Pova 6 Pro 5G", "chipset": "Dimensity 6080", "camera": "108MP Triple", "battery": "6000mAh", "base_price": 2_500_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",350_000)], "colors": ["Meteorite Grey","Comet Green","Startrail Orange"]},
    # OnePlus
    {"brand": "OnePlus", "series": "13", "chipset": "Snapdragon 8 Elite", "camera": "50MP Hasselblad Triple", "battery": "6000mAh", "base_price": 12_000_000, "ram_storage": [("12GB","256GB",0), ("16GB","512GB",2_000_000)], "colors": ["Black Eclipse","Arctic Dawn","Midnight Ocean"]},
    {"brand": "OnePlus", "series": "Nord CE4", "chipset": "Snapdragon 7 Gen 3", "camera": "50MP Dual", "battery": "5500mAh", "base_price": 4_000_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",600_000)], "colors":["Celadon Marble","Dark Chrome"]},
    # Sony
    {"brand": "Sony", "series": "Xperia 1 VI", "chipset": "Snapdragon 8 Gen 3", "camera": "52MP Zeiss Triple", "battery": "5000mAh", "base_price": 18_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",2_000_000)], "colors": ["Black","Platinum Silver","Khaki Green"]},
    {"brand": "Sony", "series": "Xperia 5 VI", "chipset": "Snapdragon 8 Gen 3", "camera": "48MP Zeiss Triple", "battery": "5000mAh", "base_price": 13_000_000, "ram_storage": [("8GB","256GB",0)], "colors": ["Black","Blue","Platinum Silver"]},
    # Honor
    {"brand": "Honor", "series": "Magic6 Pro", "chipset": "Snapdragon 8 Gen 3", "camera": "50MP Triple", "battery": "5600mAh", "base_price": 11_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_500_000)], "colors": ["Black","White","Green"]},
    {"brand": "Honor", "series": "X9b", "chipset": "Snapdragon 6 Gen 1", "camera": "108MP Triple", "battery": "5800mAh", "base_price": 3_000_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",400_000)], "colors": ["Titanium Silver","Peacock Blue","Midnight Black"]},
    # Motorola
    {"brand": "Motorola", "series": "Edge 50 Pro", "chipset": "Snapdragon 7 Gen 3", "camera": "50MP Triple", "battery": "4500mAh", "base_price": 5_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_000_000)], "colors": ["Black Beauty","Luxe Lavender","Moonlight Pearl"]},
    {"brand": "Motorola", "series": "Moto G85 5G", "chipset": "Snapdragon 6s Gen 3", "camera": "50MP Dual", "battery": "5000mAh", "base_price": 2_800_000, "ram_storage": [("8GB","128GB",0), ("12GB","256GB",500_000)], "colors": ["Cobalt Blue","Olive Green","Urban Grey"]},
    # Nokia
    {"brand": "Nokia", "series": "G42 5G", "chipset": "Snapdragon 480+", "camera": "50MP Triple", "battery": "4900mAh", "base_price": 1_900_000, "ram_storage": [("6GB","128GB",0)], "colors": ["So Pink","So Blue","So Grey"]},
]

# ─── LAPTOPS ──────────────────────────────────────────────────────────────────
LAPTOP_MODELS = [
    # MacBook
    {"brand": "Apple", "series": "MacBook Pro 16\"", "processor": "Apple M4 Max", "display": "16.2-inch Liquid Retina XDR", "battery": "Up to 22 hours", "base_price": 52_000_000, "ram_storage": [("36GB","1TB",0), ("48GB","1TB",7_000_000), ("48GB","2TB",10_000_000)], "colors": ["Space Black","Silver"]},
    {"brand": "Apple", "series": "MacBook Pro 14\"", "processor": "Apple M4 Pro", "display": "14.2-inch Liquid Retina XDR", "battery": "Up to 24 hours", "base_price": 32_000_000, "ram_storage": [("24GB","512GB",0), ("24GB","1TB",4_000_000), ("48GB","1TB",10_000_000)], "colors": ["Space Black","Silver"]},
    {"brand": "Apple", "series": "MacBook Air 15\"", "processor": "Apple M3", "display": "15.3-inch Liquid Retina", "battery": "Up to 18 hours", "base_price": 22_000_000, "ram_storage": [("8GB","256GB",0), ("8GB","512GB",2_500_000), ("16GB","512GB",4_000_000), ("24GB","2TB",9_000_000)], "colors": ["Midnight","Starlight","Space Gray","Silver"]},
    {"brand": "Apple", "series": "MacBook Air 13\"", "processor": "Apple M3", "display": "13.6-inch Liquid Retina", "battery": "Up to 18 hours", "base_price": 18_000_000, "ram_storage": [("8GB","256GB",0), ("8GB","512GB",2_000_000), ("16GB","512GB",3_500_000), ("24GB","2TB",8_000_000)], "colors": ["Midnight","Starlight","Space Gray","Silver"]},
    # ASUS
    {"brand": "ASUS", "series": "ROG Zephyrus G16", "processor": "Intel Core Ultra 9 185H", "display": "16-inch OLED 240Hz", "battery": "90Wh", "base_price": 35_000_000, "ram_storage": [("16GB","1TB",0), ("32GB","1TB",4_000_000), ("32GB","2TB",7_000_000)], "colors": ["Eclipse Gray","Platinum White"]},
    {"brand": "ASUS", "series": "ROG Strix G16", "processor": "Intel Core i9-14900HX", "display": "16-inch IPS 240Hz", "battery": "90Wh", "base_price": 28_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000), ("32GB","1TB",4_000_000)], "colors": ["Eclipse Gray","Volt Green"]},
    {"brand": "ASUS", "series": "VivoBook 16X OLED", "processor": "Intel Core i7-13700H", "display": "16-inch OLED 120Hz", "battery": "70Wh", "base_price": 14_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000)], "colors": ["Indie Black","Cool Silver"]},
    {"brand": "ASUS", "series": "ZenBook 14 OLED", "processor": "Intel Core Ultra 7 155H", "display": "14-inch OLED 120Hz", "battery": "75Wh", "base_price": 16_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",4_000_000)], "colors": ["Jasper Gray","Ponder Blue"]},
    {"brand": "ASUS", "series": "TUF Gaming A15", "processor": "AMD Ryzen 9 7940HX", "display": "15.6-inch FHD 144Hz", "battery": "90Wh", "base_price": 15_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000), ("32GB","1TB",3_500_000)], "colors": ["Mecha Gray","Jaeger Gray"]},
    {"brand": "ASUS", "series": "ExpertBook B9 OLED", "processor": "Intel Core Ultra 7 165U", "display": "14-inch OLED", "battery": "63Wh", "base_price": 22_000_000, "ram_storage": [("16GB","1TB",0), ("32GB","1TB",4_000_000)], "colors": ["Star Black"]},
    # Lenovo
    {"brand": "Lenovo", "series": "ThinkPad X1 Carbon Gen 12", "processor": "Intel Core Ultra 7 165U", "display": "14-inch IPS 2.8K OLED", "battery": "57Wh", "base_price": 32_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",5_000_000), ("64GB","2TB",12_000_000)], "colors": ["Deep Black"]},
    {"brand": "Lenovo", "series": "IdeaPad Slim 5 Gen 9", "processor": "Intel Core Ultra 5 125U", "display": "16-inch IPS 120Hz", "battery": "75Wh", "base_price": 11_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",1_500_000)], "colors": ["Abyss Blue","Cloud Grey"]},
    {"brand": "Lenovo", "series": "LOQ 15IRX9", "processor": "Intel Core i7-13650HX", "display": "15.6-inch IPS 144Hz", "battery": "60Wh", "base_price": 14_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000), ("32GB","1TB",3_500_000)], "colors": ["Luna Gray","Phantom Blue"]},
    {"brand": "Lenovo", "series": "Legion 5i Gen 9", "processor": "Intel Core i7-14650HX", "display": "15.6-inch IPS 165Hz", "battery": "80Wh", "base_price": 20_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",4_000_000), ("32GB","2TB",6_000_000)], "colors": ["Luna Gray","Glacier White"]},
    {"brand": "Lenovo", "series": "Yoga 9i Gen 9", "processor": "Intel Core Ultra 9 185H", "display": "14-inch OLED 120Hz Touch", "battery": "75Wh", "base_price": 25_000_000, "ram_storage": [("32GB","1TB",0), ("32GB","2TB",3_000_000)], "colors": ["Oatmeal","Tidal Teal"]},
    # HP
    {"brand": "HP", "series": "Spectre x360 14", "processor": "Intel Core Ultra 7 155H", "display": "14-inch OLED Touch 2.8K", "battery": "72Wh", "base_price": 23_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",5_000_000)], "colors": ["Nightfall Black","Nocturne Blue"]},
    {"brand": "HP", "series": "Envy 16", "processor": "Intel Core Ultra 7 155H", "display": "16-inch OLED 120Hz", "battery": "83Wh", "base_price": 18_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",4_000_000)], "colors": ["Meteor Silver","Nightfall Black"]},
    {"brand": "HP", "series": "OMEN 16", "processor": "Intel Core i7-13700HX", "display": "16.1-inch IPS 165Hz", "battery": "83Wh", "base_price": 19_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000), ("32GB","1TB",4_000_000)], "colors": ["Shadow Black"]},
    {"brand": "HP", "series": "Pavilion 15", "processor": "AMD Ryzen 5 7530U", "display": "15.6-inch FHD IPS", "battery": "43Wh", "base_price": 8_000_000, "ram_storage": [("8GB","512GB",0), ("16GB","512GB",1_000_000), ("16GB","1TB",2_000_000)], "colors": ["Natural Silver","Warm Gold"]},
    # Dell
    {"brand": "Dell", "series": "XPS 15 9530", "processor": "Intel Core i9-13900H", "display": "15.6-inch OLED 3.5K Touch", "battery": "86Wh", "base_price": 32_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",5_000_000), ("64GB","2TB",12_000_000)], "colors": ["Platinum Silver","Graphite"]},
    {"brand": "Dell", "series": "XPS 13 9340", "processor": "Intel Core Ultra 7 155H", "display": "13.4-inch FHD+ IPS", "battery": "55Wh", "base_price": 22_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",5_000_000)], "colors": ["Platinum","Graphite"]},
    {"brand": "Dell", "series": "Inspiron 15 3535", "processor": "AMD Ryzen 5 7530U", "display": "15.6-inch FHD IPS", "battery": "41Wh", "base_price": 7_500_000, "ram_storage": [("8GB","512GB",0), ("16GB","512GB",1_000_000)], "colors": ["Carbon Black","Platinum Silver"]},
    {"brand": "Dell", "series": "Alienware m18 R2", "processor": "Intel Core i9-14900HX", "display": "18-inch QHD+ 165Hz", "battery": "99.9Wh", "base_price": 45_000_000, "ram_storage": [("32GB","1TB",0), ("64GB","2TB",10_000_000)], "colors": ["Dark Metallic Moon"]},
    # Acer
    {"brand": "Acer", "series": "Predator Helios 16", "processor": "Intel Core i9-14900HX", "display": "16-inch IPS 240Hz", "battery": "90Wh", "base_price": 28_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",4_000_000), ("32GB","2TB",7_000_000)], "colors": ["Abyssal Black"]},
    {"brand": "Acer", "series": "Swift 14 AI", "processor": "Intel Core Ultra 5 125U", "display": "14-inch IPS OLED", "battery": "65Wh", "base_price": 13_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000)], "colors": ["Stellar Silver","Steel Gray"]},
    {"brand": "Acer", "series": "Aspire 5 A515", "processor": "AMD Ryzen 5 7520U", "display": "15.6-inch FHD IPS", "battery": "50Wh", "base_price": 7_000_000, "ram_storage": [("8GB","512GB",0), ("16GB","512GB",900_000), ("16GB","1TB",1_800_000)], "colors": ["Steel Gray","Pure Silver"]},
    {"brand": "Acer", "series": "Nitro 16 AN16", "processor": "AMD Ryzen 7 7745HX", "display": "16-inch IPS 165Hz", "battery": "90Wh", "base_price": 16_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000), ("32GB","1TB",3_500_000)], "colors": ["Obsidian Black","Shale Black"]},
    # MSI
    {"brand": "MSI", "series": "Titan GT77 HX", "processor": "Intel Core i9-13980HX", "display": "17.3-inch UHD 144Hz Mini-LED", "battery": "99.9Wh", "base_price": 55_000_000, "ram_storage": [("32GB","2TB",0), ("64GB","4TB",15_000_000)], "colors": ["Core Black"]},
    {"brand": "MSI", "series": "Raider GE78 HX", "processor": "Intel Core i9-14900HX", "display": "17-inch QHD+ 240Hz Mini-LED", "battery": "99.9Wh", "base_price": 40_000_000, "ram_storage": [("32GB","1TB",0), ("64GB","2TB",8_000_000)], "colors": ["Core Black"]},
    {"brand": "MSI", "series": "Stealth 16 AI Studio", "processor": "Intel Core Ultra 9 185H", "display": "16-inch UHD+ OLED 120Hz", "battery": "99.9Wh", "base_price": 32_000_000, "ram_storage": [("32GB","1TB",0), ("32GB","2TB",5_000_000)], "colors": ["Nebula Blue","Stealth Black"]},
    {"brand": "MSI", "series": "Prestige 13 AI Evo", "processor": "Intel Core Ultra 7 155U", "display": "13.3-inch FHD+ IPS", "battery": "72Wh", "base_price": 14_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",3_500_000)], "colors": ["Stellar Gray","Pure White"]},
    # Gigabyte
    {"brand": "Gigabyte", "series": "AORUS 16X AKG", "processor": "Intel Core i9-14900HX", "display": "16-inch QHD 240Hz", "battery": "99Wh", "base_price": 30_000_000, "ram_storage": [("32GB","1TB",0), ("32GB","2TB",5_000_000)], "colors": ["Black"]},
    {"brand": "Gigabyte", "series": "G6X 9KG", "processor": "Intel Core i7-13650HX", "display": "16-inch FHD 165Hz", "battery": "99Wh", "base_price": 17_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",2_000_000), ("32GB","1TB",3_500_000)], "colors": ["Black"]},
    # Xiaomi Laptop
    {"brand": "Xiaomi", "series": "Book Pro 14 2024", "processor": "Intel Core Ultra 7 155H", "display": "14-inch 3K OLED 120Hz", "battery": "75Wh", "base_price": 13_500_000, "ram_storage": [("16GB","1TB",0), ("32GB","1TB",3_000_000)], "colors": ["Moonlight Silver","Titanium Gray"]},
    {"brand": "Xiaomi", "series": "Book Air 13 2024", "processor": "Intel Core Ultra 5 125U", "display": "13.3-inch IPS 2.8K", "battery": "56Wh", "base_price": 9_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",1_500_000)], "colors": ["Silver","Gray"]},
]

# ─── TABLETS ──────────────────────────────────────────────────────────────────
TABLET_MODELS = [
    # iPad
    {"brand": "Apple", "series": "iPad Pro 13-inch M4", "chipset": "Apple M4", "display": "13-inch Ultra Retina XDR OLED", "battery": "38.99Wh", "base_price": 22_000_000, "ram_storage": [("8GB","256GB",0), ("8GB","512GB",3_000_000), ("16GB","1TB",7_000_000), ("16GB","2TB",11_000_000)], "connectivity": ["Wi-Fi","Wi-Fi + Cellular"]},
    {"brand": "Apple", "series": "iPad Pro 11-inch M4", "chipset": "Apple M4", "display": "11-inch Ultra Retina XDR OLED", "battery": "31.29Wh", "base_price": 16_000_000, "ram_storage": [("8GB","256GB",0), ("8GB","512GB",3_000_000), ("16GB","1TB",7_000_000)], "connectivity": ["Wi-Fi","Wi-Fi + Cellular"]},
    {"brand": "Apple", "series": "iPad Air 13-inch M2", "chipset": "Apple M2", "display": "13-inch Liquid Retina", "battery": "36.59Wh", "base_price": 14_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",2_000_000), ("8GB","512GB",4_000_000), ("8GB","1TB",7_000_000)], "connectivity": ["Wi-Fi","Wi-Fi + Cellular"]},
    {"brand": "Apple", "series": "iPad Air 11-inch M2", "chipset": "Apple M2", "display": "11-inch Liquid Retina", "battery": "28.65Wh", "base_price": 11_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",2_000_000), ("8GB","512GB",4_000_000)], "connectivity": ["Wi-Fi","Wi-Fi + Cellular"]},
    {"brand": "Apple", "series": "iPad 10th Gen", "chipset": "Apple A14 Bionic", "display": "10.9-inch Liquid Retina", "battery": "28.65Wh", "base_price": 7_000_000, "ram_storage": [("4GB","64GB",0), ("4GB","256GB",2_000_000)], "connectivity": ["Wi-Fi","Wi-Fi + Cellular"]},
    {"brand": "Apple", "series": "iPad mini 7", "chipset": "Apple A17 Pro", "display": "8.3-inch Liquid Retina", "battery": "19.32Wh", "base_price": 9_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",2_000_000), ("8GB","512GB",4_000_000)], "connectivity": ["Wi-Fi","Wi-Fi + Cellular"]},
    # Samsung Tab
    {"brand": "Samsung", "series": "Galaxy Tab S10 Ultra", "chipset": "Snapdragon 8 Gen 3", "display": "14.6-inch Dynamic AMOLED 2X 120Hz", "battery": "11200mAh", "base_price": 18_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",2_500_000)], "connectivity": ["Wi-Fi","Wi-Fi + LTE"]},
    {"brand": "Samsung", "series": "Galaxy Tab S10+", "chipset": "Snapdragon 8 Gen 3", "display": "12.4-inch Dynamic AMOLED 2X 120Hz", "battery": "10090mAh", "base_price": 14_500_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",2_500_000)], "connectivity": ["Wi-Fi","Wi-Fi + LTE"]},
    {"brand": "Samsung", "series": "Galaxy Tab S10", "chipset": "Snapdragon 8 Gen 3", "display": "11-inch Dynamic AMOLED 2X 120Hz", "battery": "8000mAh", "base_price": 11_000_000, "ram_storage": [("12GB","256GB",0)], "connectivity": ["Wi-Fi","Wi-Fi + LTE"]},
    {"brand": "Samsung", "series": "Galaxy Tab S9 FE", "chipset": "Exynos 1380", "display": "10.9-inch TFT LCD 90Hz", "battery": "8000mAh", "base_price": 6_500_000, "ram_storage": [("6GB","128GB",0), ("8GB","256GB",1_000_000)], "connectivity": ["Wi-Fi","Wi-Fi + LTE"]},
    {"brand": "Samsung", "series": "Galaxy Tab A9+", "chipset": "Snapdragon 695", "display": "11-inch TFT LCD 90Hz", "battery": "7040mAh", "base_price": 4_000_000, "ram_storage": [("4GB","64GB",0), ("8GB","128GB",700_000)], "connectivity": ["Wi-Fi","Wi-Fi + LTE"]},
    # Xiaomi Pad
    {"brand": "Xiaomi", "series": "Pad 6S Pro 12.4", "chipset": "Snapdragon 8 Gen 2", "display": "12.4-inch LCD 144Hz", "battery": "10000mAh", "base_price": 8_500_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",1_000_000), ("12GB","512GB",2_000_000)], "connectivity": ["Wi-Fi","Wi-Fi + 5G"]},
    {"brand": "Xiaomi", "series": "Pad 6 Pro", "chipset": "Snapdragon 8+ Gen 1", "display": "11-inch LCD 144Hz", "battery": "8600mAh", "base_price": 6_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",700_000), ("12GB","256GB",1_200_000)], "connectivity": ["Wi-Fi","Wi-Fi + 5G"]},
    {"brand": "Xiaomi", "series": "Redmi Pad Pro", "chipset": "Snapdragon 7s Gen 2", "display": "12.1-inch LCD 120Hz", "battery": "10000mAh", "base_price": 4_000_000, "ram_storage": [("6GB","128GB",0), ("8GB","256GB",600_000)], "connectivity": ["Wi-Fi","Wi-Fi + 4G"]},
    # Lenovo Tab
    {"brand": "Lenovo", "series": "Tab P12 Pro", "chipset": "Snapdragon 870", "display": "12.6-inch AMOLED 120Hz", "battery": "10200mAh", "base_price": 8_000_000, "ram_storage": [("8GB","256GB",0)], "connectivity": ["Wi-Fi","Wi-Fi + 5G"]},
    {"brand": "Lenovo", "series": "Tab P11 Gen 2", "chipset": "Helio G99", "display": "11.5-inch LCD 120Hz", "battery": "7700mAh", "base_price": 3_500_000, "ram_storage": [("6GB","128GB",0), ("6GB","256GB",500_000)], "connectivity": ["Wi-Fi","Wi-Fi + LTE"]},
    # OPPO Pad
    {"brand": "OPPO", "series": "Pad 2", "chipset": "Dimensity 9000", "display": "12.1-inch IPS LCD 144Hz", "battery": "9510mAh", "base_price": 7_000_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",700_000), ("12GB","256GB",1_500_000)], "connectivity": ["Wi-Fi","Wi-Fi + 5G"]},
    {"brand": "Realme", "series": "Pad 2", "chipset": "Helio G99", "display": "11.5-inch IPS LCD 120Hz", "battery": "8360mAh", "base_price": 2_800_000, "ram_storage": [("4GB","64GB",0), ("6GB","128GB",400_000)], "connectivity": ["Wi-Fi","Wi-Fi + LTE"]},
]

# ─── SMARTWATCHES ──────────────────────────────────────────────────────────────
SMARTWATCH_MODELS = [
    # Apple Watch
    {"brand": "Apple", "series": "Apple Watch Series 10", "processor": "S10", "display": "45mm / 46mm Always-On Retina LTPO OLED", "battery": "Up to 18 hours", "base_price": 6_500_000, "variants": [("41mm","Aluminum","Sport Loop",0), ("41mm","Aluminum","Sport Band",0), ("45mm","Aluminum","Sport Band",500_000), ("45mm","Titanium","Milanese Loop",4_000_000)], "connectivity": "GPS + Cellular"},
    {"brand": "Apple", "series": "Apple Watch Ultra 2", "processor": "S9", "display": "49mm Always-On Retina LTPO OLED", "battery": "Up to 60 hours", "base_price": 13_000_000, "variants": [("49mm","Titanium","Alpine Loop",0), ("49mm","Titanium","Trail Loop",0), ("49mm","Black Titanium","Ocean Band",500_000)], "connectivity": "GPS + Cellular"},
    {"brand": "Apple", "series": "Apple Watch SE (3rd Gen)", "processor": "S9", "display": "40mm / 44mm Retina LTPO OLED", "battery": "Up to 18 hours", "base_price": 3_800_000, "variants": [("40mm","Aluminum","Sport Band",0), ("44mm","Aluminum","Sport Band",500_000)], "connectivity": "GPS / GPS + Cellular"},
    # Samsung
    {"brand": "Samsung", "series": "Galaxy Watch 7", "processor": "Exynos W1000", "display": "40mm / 44mm Super AMOLED Always-On", "battery": "300mAh / 425mAh", "base_price": 3_800_000, "variants": [("40mm","Aluminum","Green",0), ("40mm","Aluminum","Cream",0), ("44mm","Aluminum","Green",500_000), ("44mm","Aluminum","Silver",500_000)], "connectivity": "Bluetooth / LTE"},
    {"brand": "Samsung", "series": "Galaxy Watch Ultra", "processor": "Exynos W1000", "display": "47mm Super AMOLED 60Hz", "battery": "590mAh", "base_price": 7_500_000, "variants": [("47mm","Titanium White","White Marine Band",0), ("47mm","Titanium Black","Black Marine Band",0), ("47mm","Titanium Silver","Gray Marine Band",0)], "connectivity": "Bluetooth + LTE"},
    {"brand": "Samsung", "series": "Galaxy Watch FE", "processor": "Exynos W920", "display": "40mm Super AMOLED", "battery": "247mAh", "base_price": 2_200_000, "variants": [("40mm","Aluminum","Pink Gold",0), ("40mm","Aluminum","Graphite",0), ("40mm","Aluminum","Silver",0)], "connectivity": "Bluetooth / LTE"},
    # Garmin
    {"brand": "Garmin", "series": "Fenix 8 Solar", "processor": "N/A", "display": "51mm / 47mm AMOLED MIP Solar", "battery": "Up to 29 days", "base_price": 14_000_000, "variants": [("47mm","Carbon Gray","Black Band",0), ("47mm","Silver","White Band",0), ("51mm","Carbon Gray","Black Band",2_000_000), ("51mm","Silver","White Band",2_000_000)], "connectivity": "GPS"},
    {"brand": "Garmin", "series": "Forerunner 965", "processor": "N/A", "display": "47mm AMOLED", "battery": "Up to 31 days GPS", "base_price": 9_000_000, "variants": [("47mm","Carbon Gray DLC","Black Band",0), ("47mm","Whitestone","White Band",0)], "connectivity": "GPS"},
    {"brand": "Garmin", "series": "Venu 3", "processor": "N/A", "display": "41mm / 45mm AMOLED", "battery": "Up to 14 days", "base_price": 5_500_000, "variants": [("41mm","Ivory","Ivory Band",0), ("41mm","Black","Slate Band",0), ("45mm","Mist Gray","Whitestone Band",600_000), ("45mm","Black","Slate Band",600_000)], "connectivity": "GPS + Cellular"},
    {"brand": "Garmin", "series": "Instinct 3", "processor": "N/A", "display": "45mm / 50mm MIP Transflective", "battery": "Up to 24 days", "base_price": 5_000_000, "variants": [("45mm","Night Vision","Black Band",0), ("45mm","Whitestone","Whitestone Band",0), ("50mm AMOLED","Darkside","Black Band",1_000_000)], "connectivity": "GPS"},
    # Amazfit
    {"brand": "Amazfit", "series": "Balance", "processor": "N/A", "display": "46mm AMOLED 60Hz", "battery": "Up to 14 days", "base_price": 1_800_000, "variants": [("46mm","Sunset Grey","Grey Band",0), ("46mm","Olive Green","Green Band",0), ("46mm","Marble White","White Band",0)], "connectivity": "GPS + Bluetooth"},
    {"brand": "Amazfit", "series": "GTR 4", "processor": "N/A", "display": "46mm AMOLED 60Hz", "battery": "Up to 14 days", "base_price": 1_500_000, "variants": [("46mm","Rucksack Khaki","Khaki Band",0), ("46mm","Superspeed Black","Black Band",0), ("46mm","Vintage Brown Leather","Brown Band",0)], "connectivity": "GPS + Bluetooth"},
    {"brand": "Amazfit", "series": "T-Rex Ultra", "processor": "N/A", "display": "47mm AMOLED", "battery": "Up to 20 days", "base_price": 3_000_000, "variants": [("47mm","Falcon Black","Black Band",0), ("47mm","Sahara Khaki","Khaki Band",0)], "connectivity": "GPS + Dual Frequency"},
    {"brand": "Amazfit", "series": "Bip 5 Unity", "processor": "N/A", "display": "49mm TFT-LCD", "battery": "Up to 10 days", "base_price": 600_000, "variants": [("49mm","Pink","Pink Band",0), ("49mm","Blue","Blue Band",0), ("49mm","Black","Black Band",0), ("49mm","Cream","Cream Band",0)], "connectivity": "GPS + Bluetooth"},
    # Huawei Watch
    {"brand": "Huawei", "series": "Watch GT 5 Pro", "processor": "N/A", "display": "46mm / 42mm AMOLED", "battery": "Up to 21 days", "base_price": 3_500_000, "variants": [("46mm","Titanium","Titanium Band",0), ("46mm","Dark Brown","Leather Band",0), ("42mm","White","White Ceramic Band",- 400_000)], "connectivity": "GPS + Bluetooth"},
    {"brand": "Huawei", "series": "Watch GT 5", "processor": "N/A", "display": "46mm / 41mm AMOLED", "battery": "Up to 14 days", "base_price": 2_500_000, "variants": [("46mm","Black","Black Fluoroelastomer Band",0), ("46mm","Brown","Leather Band",0), ("41mm","Pink Gold","Pink Band",-400_000), ("41mm","Silver","White Band",-400_000)], "connectivity": "GPS + Bluetooth"},
    # Xiaomi Watch
    {"brand": "Xiaomi", "series": "Watch S3", "processor": "N/A", "display": "46mm AMOLED 60Hz", "battery": "Up to 15 days", "base_price": 1_600_000, "variants": [("46mm","Black","Black Band",0), ("46mm","Silver","Silver Band",0), ("46mm","Brown","Brown Leather Band",0)], "connectivity": "GPS + Bluetooth"},
    {"brand": "Xiaomi", "series": "Redmi Watch 5 Active", "processor": "N/A", "display": "47.8mm LCD 60Hz", "battery": "Up to 18 days", "base_price": 400_000, "variants": [("47.8mm","Black","Black Band",0), ("47.8mm","Blue","Blue Band",0), ("47.8mm","Pink","Pink Band",0)], "connectivity": "GPS + Bluetooth"},
    # OPPO Watch
    {"brand": "OPPO", "series": "Watch X", "processor": "Snapdragon W5 Gen 1", "display": "46mm AMOLED 60Hz", "battery": "Up to 12 days", "base_price": 3_800_000, "variants": [("46mm","Graphite Black","Black Band",0), ("46mm","Platinum Silver","Silver Band",0)], "connectivity": "GPS + LTE"},
    # Vivo Watch
    {"brand": "Vivo", "series": "Watch 3", "processor": "N/A", "display": "46mm AMOLED", "battery": "Up to 16 days", "base_price": 2_200_000, "variants": [("46mm","Dark Knight","Black Band",0), ("46mm","Silver Wave","Silver Band",0)], "connectivity": "GPS + Bluetooth"},
]


def make_sku(prefix, idx):
    return f"{prefix}-{idx:05d}"


rows = []  # list of dicts

# ─── Generate Smartphone rows ──────────────────────────────────────────────────
sku_idx = 1
for m in SMARTPHONE_MODELS:
    for ram, storage, price_delta in m["ram_storage"]:
        for color in m["colors"]:
            price = m["base_price"] + price_delta
            rows.append({
                "SKU": make_sku("HP", sku_idx),
                "Nama Produk": m["series"],
                "Kategori": "Smartphone",
                "Harga": price,
                "Deskripsi": f"{m['series']} {ram}/{storage} {color} - {m['chipset']} | {m['camera']} | {m['battery']} | 5G",
                "Status Stok": random.choice(["in_stock","in_stock","in_stock","out_of_stock","preorder"]),
                "Stok": random.randint(0, 200),
                "Brand": m["brand"],
                "RAM": ram,
                "Storage": storage,
                "Warna": color,
                "Chipset": m["chipset"],
                "Kamera": m["camera"],
                "Baterai": m["battery"],
            })
            sku_idx += 1

# ─── Generate Laptop rows ──────────────────────────────────────────────────────
for m in LAPTOP_MODELS:
    for ram, storage, price_delta in m["ram_storage"]:
        for color in m["colors"]:
            price = m["base_price"] + price_delta
            rows.append({
                "SKU": make_sku("LT", sku_idx),
                "Nama Produk": m["series"],
                "Kategori": "Laptop",
                "Harga": price,
                "Deskripsi": f"{m['series']} {ram}/{storage} {color} - {m['processor']} | {m['display']} | Baterai {m['battery']}",
                "Status Stok": random.choice(["in_stock","in_stock","in_stock","out_of_stock","preorder"]),
                "Stok": random.randint(0, 50),
                "Brand": m["brand"],
                "RAM": ram,
                "Storage": storage,
                "Warna": color,
                "Prosesor": m["processor"],
                "Layar": m["display"],
                "Baterai": m["battery"],
            })
            sku_idx += 1

# ─── Generate Tablet rows ──────────────────────────────────────────────────────
for m in TABLET_MODELS:
    for ram, storage, price_delta in m["ram_storage"]:
        for conn in m.get("connectivity", ["Wi-Fi"]):
            price = m["base_price"] + price_delta + (1_500_000 if "Cellular" in conn or "LTE" in conn or "5G" in conn else 0)
            rows.append({
                "SKU": make_sku("TB", sku_idx),
                "Nama Produk": m["series"],
                "Kategori": "Tablet",
                "Harga": price,
                "Deskripsi": f"{m['series']} {ram}/{storage} {conn} - {m['chipset']} | {m['display']} | {m['battery']}",
                "Status Stok": random.choice(["in_stock","in_stock","out_of_stock","preorder"]),
                "Stok": random.randint(0, 80),
                "Brand": m["brand"],
                "RAM": ram,
                "Storage": storage,
                "Konektivitas": conn,
                "Chipset": m["chipset"],
                "Layar": m["display"],
                "Baterai": m["battery"],
            })
            sku_idx += 1

# ─── Generate Smartwatch rows ─────────────────────────────────────────────────
for m in SMARTWATCH_MODELS:
    for variant in m["variants"]:
        size, material, band, price_delta = variant
        price = m["base_price"] + price_delta
        rows.append({
            "SKU": make_sku("SW", sku_idx),
            "Nama Produk": m["series"],
            "Kategori": "Smartwatch",
            "Harga": price,
            "Deskripsi": f"{m['series']} {size} {material} - {m['display']} | Baterai {m['battery']} | {m['connectivity']}",
            "Status Stok": random.choice(["in_stock","in_stock","in_stock","out_of_stock"]),
            "Stok": random.randint(0, 100),
            "Brand": m["brand"],
            "Ukuran": size,
            "Material": material,
            "Tali": band,
            "Layar": m["display"],
            "Baterai": m["battery"],
            "Konektivitas": m["connectivity"],
        })
        sku_idx += 1

print(f"Generated {len(rows)} rows before padding")

# ─── Pad to 5000 rows with realistic variants ─────────────────────────────────
EXTRA_PHONES = [
    # Budget Android brands
    {"brand": "Infinix", "series": "Note 40", "chipset": "Helio G99", "camera": "108MP Triple", "battery": "5000mAh", "base_price": 2_300_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",400_000)], "colors": ["Titan Gold","Obsidian Black","Racing Green"]},
    {"brand": "Infinix", "series": "Smart 8 Pro", "chipset": "Helio G36", "camera": "50MP Dual", "battery": "5000mAh", "base_price": 1_200_000, "ram_storage": [("4GB","64GB",0), ("8GB","128GB",300_000)], "colors": ["Timber Black","Galaxy White","Bora Purple"]},
    {"brand": "Tecno", "series": "Spark 30 Pro", "chipset": "Helio G88", "camera": "64MP Triple", "battery": "5000mAh", "base_price": 2_000_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",350_000)], "colors": ["Amber Gold","Camo Green","Starfall Silver"]},
    {"brand": "Tecno", "series": "Camon 30 Premier 5G", "chipset": "Dimensity 8200", "camera": "50MP Triple ZEISS", "battery": "5000mAh", "base_price": 5_000_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",600_000)], "colors": ["Space Black","Champagne Gold"]},
    {"brand": "Vivo", "series": "Y18", "chipset": "Helio G85", "camera": "50MP Dual", "battery": "5000mAh", "base_price": 1_500_000, "ram_storage": [("4GB","128GB",0), ("6GB","128GB",200_000)], "colors":["Space Black","Gem Green","Glitter Gold"]},
    {"brand": "Vivo", "series": "T3 Pro 5G", "chipset": "Snapdragon 7 Gen 3", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 3_800_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",400_000), ("12GB","256GB",800_000)], "colors": ["Sparkle Green","Midnight Blue","Titanium Gold"]},
    {"brand": "Realme", "series": "13 Pro+ 5G", "chipset": "Snapdragon 7s Gen 2", "camera": "50MP Periscope Triple", "battery": "5200mAh", "base_price": 5_500_000, "ram_storage": [("8GB","256GB",0), ("12GB","256GB",600_000), ("12GB","512GB",1_200_000)], "colors": ["Emerald Green","Monet Purple","Fluid Silver"]},
    {"brand": "OPPO", "series": "A3", "chipset": "Snapdragon 4 Gen 2", "camera": "50MP Dual", "battery": "5100mAh", "base_price": 2_300_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",350_000)], "colors": ["Starry Black","Dreamy Purple","Glowwave Blue"]},
    {"brand": "Xiaomi", "series": "14", "chipset": "Snapdragon 8 Gen 3", "camera": "50MP Leica Triple", "battery": "4610mAh", "base_price": 12_500_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_500_000), ("16GB","512GB",2_500_000)], "colors": ["Black","White","Jade Green","Pink"]},
    {"brand": "Samsung", "series": "Galaxy F55 5G", "chipset": "Snapdragon 7 Gen 1", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 4_500_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",500_000), ("12GB","256GB",1_000_000)], "colors": ["Apricot Crush","Meteor Violet","Peacock Blue"]},
    {"brand": "OnePlus", "series": "12R", "chipset": "Snapdragon 8 Gen 1", "camera": "50MP Triple", "battery": "5500mAh", "base_price": 6_000_000, "ram_storage": [("8GB","128GB",0), ("16GB","256GB",1_500_000)], "colors": ["Iron Gray","Cool Blue"]},
    {"brand": "Motorola", "series": "Razr 50 Ultra", "chipset": "Snapdragon 8s Gen 3", "camera": "50MP Dual", "battery": "4000mAh", "base_price": 15_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_500_000)], "colors": ["Peach Fuzz","Spring Bud","Midnight Blue"]},
    {"brand": "Nothing", "series": "Phone (2a)", "chipset": "Dimensity 7200 Pro", "camera": "50MP Dual", "battery": "5000mAh", "base_price": 4_500_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",500_000), ("12GB","256GB",900_000)], "colors": ["Black","White","Blue Special Edition"]},
    {"brand": "Nothing", "series": "Phone (2)", "chipset": "Snapdragon 8+ Gen 1", "camera": "50MP Dual", "battery": "4700mAh", "base_price": 7_000_000, "ram_storage": [("8GB","128GB",0), ("12GB","256GB",1_500_000)], "colors": ["Dark Grey","White"]},
    {"brand": "Huawei", "series": "Pura 70 Pro", "chipset": "Kirin 9010", "camera": "50MP Triple", "battery": "5050mAh", "base_price": 14_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",2_000_000)], "colors": ["Black","White","Brown"]},
    {"brand": "Honor", "series": "200 Pro", "chipset": "Snapdragon 8s Gen 3", "camera": "50MP Periscope Triple", "battery": "5200mAh", "base_price": 8_000_000, "ram_storage": [("12GB","256GB",0), ("12GB","512GB",1_500_000)], "colors": ["Midnight Black","Moonlight White","Violet"]},
    {"brand": "Asus", "series": "ROG Phone 8 Pro", "chipset": "Snapdragon 8 Gen 3", "camera": "50MP Triple", "battery": "5500mAh", "base_price": 18_000_000, "ram_storage": [("16GB","512GB",0), ("24GB","1TB",5_000_000)], "colors": ["Phantom Black","Storm White"]},
    {"brand": "Asus", "series": "Zenfone 11 Ultra", "chipset": "Snapdragon 8 Gen 3", "camera": "50MP Triple", "battery": "5500mAh", "base_price": 13_000_000, "ram_storage": [("12GB","256GB",0), ("16GB","512GB",2_500_000)], "colors": ["Skyline Blue","Eternal Black","Misty Gray","Desert Sand"]},
    {"brand": "Google", "series": "Pixel 9 Pro XL", "chipset": "Google Tensor G4", "camera": "50MP Triple", "battery": "5060mAh", "base_price": 16_000_000, "ram_storage": [("16GB","128GB",0), ("16GB","256GB",1_500_000), ("16GB","512GB",3_000_000), ("16GB","1TB",6_000_000)], "colors": ["Obsidian","Porcelain","Hazel","Rose Quartz"]},
    {"brand": "Google", "series": "Pixel 9", "chipset": "Google Tensor G4", "camera": "50MP Dual", "battery": "4700mAh", "base_price": 11_000_000, "ram_storage": [("12GB","128GB",0), ("12GB","256GB",1_200_000)], "colors": ["Obsidian","Porcelain","Wintergreen","Peony"]},
    {"brand": "Xiaomi", "series": "POCO F6 Pro", "chipset": "Snapdragon 8 Gen 2", "camera": "50MP Triple", "battery": "5000mAh", "base_price": 7_000_000, "ram_storage": [("12GB","256GB",0), ("16GB","512GB",1_500_000), ("16GB","1TB",3_000_000)], "colors": ["Black","White","Green"]},
]

extra_rows = []
for m in EXTRA_PHONES:
    for ram, storage, price_delta in m["ram_storage"]:
        for color in m["colors"]:
            price = m["base_price"] + price_delta
            extra_rows.append({
                "SKU": make_sku("HP", sku_idx),
                "Nama Produk": m["series"],
                "Kategori": "Smartphone",
                "Harga": price,
                "Deskripsi": f"{m['series']} {ram}/{storage} {color} - {m['chipset']} | {m['camera']} | {m['battery']}",
                "Status Stok": random.choice(["in_stock","in_stock","in_stock","out_of_stock","preorder"]),
                "Stok": random.randint(0, 200),
                "Brand": m["brand"],
                "RAM": ram,
                "Storage": storage,
                "Warna": color,
                "Chipset": m["chipset"],
                "Kamera": m["camera"],
                "Baterai": m["battery"],
            })
            sku_idx += 1

rows.extend(extra_rows)

# ─── Pad further if still < 5000 ─────────────────────────────────────────────
EXTRA_LAPTOPS = [
    {"brand": "ASUS", "series": "ROG Flow Z13 2024", "processor": "Intel Core Ultra 9 185H", "display": "13.4-inch QHD+ 180Hz Touch", "battery": "38Wh", "base_price": 30_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",5_000_000)], "colors": ["Black"]},
    {"brand": "Lenovo", "series": "ThinkPad T14s Gen 5", "processor": "Intel Core Ultra 7 155U", "display": "14-inch IPS 2.8K OLED", "battery": "58Wh", "base_price": 19_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",4_000_000)], "colors": ["Deep Black"]},
    {"brand": "HP", "series": "EliteBook 845 G11", "processor": "AMD Ryzen 5 Pro 7535U", "display": "14-inch IPS WUXGA", "battery": "51Wh", "base_price": 16_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",3_500_000)], "colors": ["Silver"]},
    {"brand": "Dell", "series": "Latitude 7450", "processor": "Intel Core Ultra 7 165U", "display": "14-inch FHD+ IPS", "battery": "65Wh", "base_price": 22_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",4_000_000)], "colors": ["Carbon Fiber Black"]},
    {"brand": "Acer", "series": "Chromebook Plus 514", "processor": "Intel Core i3-1315U", "display": "14-inch IPS FHD", "battery": "65Wh", "base_price": 5_500_000, "ram_storage": [("8GB","128GB",0), ("8GB","256GB",700_000)], "colors": ["Steel Gray"]},
    {"brand": "Microsoft", "series": "Surface Pro 11", "processor": "Snapdragon X Elite", "display": "13-inch PixelSense Flow 120Hz Touch", "battery": "Up to 14 hours", "base_price": 22_000_000, "ram_storage": [("16GB","256GB",0), ("32GB","512GB",5_000_000), ("64GB","1TB",12_000_000)], "colors": ["Platinum","Sapphire"]},
    {"brand": "Microsoft", "series": "Surface Laptop 7", "processor": "Snapdragon X Plus", "display": "13.8-inch PixelSense 120Hz Touch", "battery": "Up to 22 hours", "base_price": 18_000_000, "ram_storage": [("16GB","256GB",0), ("16GB","512GB",2_500_000), ("32GB","1TB",6_000_000)], "colors": ["Platinum","Dune","Sapphire","Black"]},
    {"brand": "Huawei", "series": "MateBook X Pro 2024", "processor": "Intel Core Ultra 9 185H", "display": "14.2-inch OLED 2.5K 120Hz Touch", "battery": "70Wh", "base_price": 23_000_000, "ram_storage": [("16GB","1TB",0), ("32GB","1TB",4_000_000)], "colors": ["Space Gray","Chalk White"]},
    {"brand": "Huawei", "series": "MateBook D16 2024", "processor": "Intel Core Ultra 5 125H", "display": "16-inch IPS FHD", "battery": "65Wh", "base_price": 10_000_000, "ram_storage": [("16GB","512GB",0), ("16GB","1TB",1_500_000)], "colors": ["Space Gray","Mystic Silver"]},
    {"brand": "Honor", "series": "MagicBook Art 14", "processor": "Intel Core Ultra 7 155H", "display": "14.5-inch OLED 120Hz", "battery": "75Wh", "base_price": 14_000_000, "ram_storage": [("16GB","1TB",0), ("32GB","1TB",3_000_000)], "colors": ["Space Gray"]},
    {"brand": "Razer", "series": "Blade 15 2024", "processor": "Intel Core i9-14900HX", "display": "15.6-inch OLED QHD 240Hz", "battery": "80Wh", "base_price": 38_000_000, "ram_storage": [("16GB","1TB",0), ("32GB","2TB",8_000_000)], "colors": ["Black"]},
    {"brand": "LG", "series": "Gram 16 2024", "processor": "Intel Core Ultra 7 155H", "display": "16-inch IPS WQXGA", "battery": "99Wh", "base_price": 20_000_000, "ram_storage": [("16GB","512GB",0), ("32GB","1TB",4_000_000)], "colors": ["White","Charcoal Gray"]},
]
for m in EXTRA_LAPTOPS:
    for ram, storage, price_delta in m["ram_storage"]:
        for color in m["colors"]:
            price = m["base_price"] + price_delta
            rows.append({
                "SKU": make_sku("LT", sku_idx),
                "Nama Produk": m["series"],
                "Kategori": "Laptop",
                "Harga": price,
                "Deskripsi": f"{m['series']} {ram}/{storage} {color} - {m['processor']} | {m['display']} | {m['battery']}",
                "Status Stok": random.choice(["in_stock","in_stock","in_stock","out_of_stock","preorder"]),
                "Stok": random.randint(0, 40),
                "Brand": m["brand"],
                "RAM": ram,
                "Storage": storage,
                "Warna": color,
                "Prosesor": m["processor"],
                "Layar": m["display"],
                "Baterai": m["battery"],
            })
            sku_idx += 1

print(f"Generated {len(rows)} rows after all brands")

# If still under 5000, duplicate some variants with slight price variations
target = 5000
if len(rows) < target:
    needed = target - len(rows)
    base_pool = [r for r in rows if r["Kategori"] == "Smartphone"]
    random.shuffle(base_pool)
    for i in range(needed):
        src = base_pool[i % len(base_pool)].copy()
        src["SKU"] = make_sku("HP", sku_idx)
        src["Harga"] = int(src["Harga"] * random.uniform(0.97, 1.03))  # ±3% price variant
        src["Stok"] = random.randint(0, 150)
        rows.append(src)
        sku_idx += 1

rows = rows[:5000]
print(f"Final row count: {len(rows)}")

# ─── Build XLS ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Katalog Produk"

HEADERS = ["SKU","Nama Produk","Kategori","Harga","Deskripsi","Status Stok","Stok",
           "Brand","RAM","Storage","Warna","Chipset","Kamera","Baterai","Prosesor","Layar","Konektivitas","Material","Ukuran","Tali"]

# Header styling
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

for col_idx, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Data rows
for row_idx, row in enumerate(rows, 2):
    for col_idx, h in enumerate(HEADERS, 1):
        ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

# Column widths
col_widths = [14, 36, 14, 14, 70, 14, 8, 14, 8, 10, 22, 24, 22, 14, 30, 32, 20, 16, 12, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = "A2"

out_path = "/data/www/aimin/public/uploads/katalog_produk_sample.xlsx"
wb.save(out_path)
print(f"Saved to {out_path}")
